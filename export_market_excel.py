import pandas as pd
import gspread
import json
import os
from google.oauth2.service_account import Credentials
from vnstock import Vnstock
from datetime import datetime, timedelta
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import yfinance as yf

OUTPUT_FILE = "market_data.xlsx"
EXCHANGES = ["HOSE", "HNX", "UPCOM"]
YF_SUFFIX = {"HOSE": ".VN", "HNX": ".HN", "UPCOM": ".VN"}

# ─────────────────────────────────────────────────────────
# GOOGLE SHEET
# ─────────────────────────────────────────────────────────
def get_gsheet_client():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    creds_dict = json.loads(creds_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.Client(auth=creds)

def push_to_gsheet(all_data: dict):
    client = get_gsheet_client()
    spreadsheet_id = os.environ.get("SPREADSHEET_ID")
    wb = client.open_by_key(spreadsheet_id)

    HEADER_FMT = {
        "backgroundColor": {"red": 0.12, "green": 0.31, "blue": 0.47},
        "textFormat": {
            "bold": True,
            "foregroundColor": {"red": 1, "green": 1, "blue": 1}
        },
        "horizontalAlignment": "CENTER"
    }

    for exchange, df in all_data.items():
        try:
            ws = wb.worksheet(exchange)
            ws.clear()
        except gspread.WorksheetNotFound:
            ws = wb.add_worksheet(title=exchange, rows=2000, cols=10)

        if df.empty:
            print(f"  ⚠️ {exchange}: không có dữ liệu")
            continue

        data = [df.columns.tolist()] + df.values.tolist()
        ws.update(data, value_input_option="RAW")
        ws.format(f"A1:{get_column_letter(len(df.columns))}1", HEADER_FMT)
        print(f"  ✅ Sheet '{exchange}': {len(df)} dòng")

    try:
        ws_all = wb.worksheet("Tổng Hợp")
        ws_all.clear()
    except gspread.WorksheetNotFound:
        ws_all = wb.add_worksheet(title="Tổng Hợp", rows=5000, cols=10)

    frames = []
    for exchange, df in all_data.items():
        if not df.empty:
            df_copy = df.copy()
            df_copy.insert(0, "Sàn", exchange)
            frames.append(df_copy)

    if frames:
        df_total = pd.concat(frames, ignore_index=True)
        data_all = [df_total.columns.tolist()] + df_total.values.tolist()
        ws_all.update(data_all, value_input_option="RAW")
        ws_all.format(f"A1:{get_column_letter(len(df_total.columns))}1", HEADER_FMT)
        print(f"  ✅ Sheet 'Tổng Hợp': {len(df_total)} dòng")

# ─────────────────────────────────────────────────────────
# EXCEL (.xlsx)
# ─────────────────────────────────────────────────────────
COL_WIDTHS = [10, 14, 14, 14, 14, 14, 16, 14, 10]

def style_sheet(ws, nrows, ncols):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_even = PatternFill("solid", fgColor="DEEAF1")
    for row_idx in range(2, nrows + 2):
        for cell in ws[row_idx]:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = fill_even
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")

        pct_cell = ws.cell(row=row_idx, column=8)
        try:
            val = float(pct_cell.value)
            pct_cell.font = Font(
                name="Arial", size=10, bold=True,
                color="00B050" if val >= 0 else "FF0000"
            )
        except (TypeError, ValueError):
            pass

    for i, w in enumerate(COL_WIDTHS[:ncols], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def export_xlsx(all_data: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    frames = []
    for exchange, df in all_data.items():
        if not df.empty:
            df_copy = df.copy()
            df_copy.insert(0, "Sàn", exchange)
            frames.append(df_copy)

    if frames:
        df_total = pd.concat(frames, ignore_index=True)
        ws_all = wb.create_sheet("Tổng Hợp", 0)
        ws_all.append(df_total.columns.tolist())
        for _, r in df_total.iterrows():
            ws_all.append(r.tolist())
        style_sheet(ws_all, len(df_total), len(df_total.columns))

    for exchange, df in all_data.items():
        ws = wb.create_sheet(exchange)
        if df.empty:
            ws.append(["Không có dữ liệu"])
            continue
        ws.append(df.columns.tolist())
        for _, r in df.iterrows():
            ws.append(r.tolist())
        style_sheet(ws, len(df), len(df.columns))

    wb.save(OUTPUT_FILE)
    print(f"  ✅ Đã lưu {OUTPUT_FILE}")

# ─────────────────────────────────────────────────────────
# LẤY DANH SÁCH MÃ
# ─────────────────────────────────────────────────────────
def get_tickers_by_exchange():
    result = {ex: [] for ex in EXCHANGES}
    try:
        stock = Vnstock().stock(symbol='ACB', source='VCI')
        df_all = stock.listing.all_symbols()
        df_all.columns = [c.lower().strip() for c in df_all.columns]

        ex_col     = next((c for c in ['exchange', 'comgroupcode', 'organ_type', 'group_code']
                           if c in df_all.columns), None)
        ticker_col = next((c for c in ['ticker', 'symbol', 'code']
                           if c in df_all.columns), None)

        if ex_col and ticker_col:
            for ex in EXCHANGES:
                mask = df_all[ex_col].str.upper().str.contains(ex)
                result[ex] = df_all[mask][ticker_col].tolist()
                print(f"  {ex}: {len(result[ex])} mã (VCI)")
        elif ticker_col:
            result["HOSE"] = df_all[ticker_col].tolist()
            print(f"  HOSE (tất cả): {len(result['HOSE'])} mã")
    except Exception as e:
        print(f"  ⚠️ VCI listing lỗi: {e}")

    return result

# ─────────────────────────────────────────────────────────
# FALLBACK CHAIN: VCI → KB → DNSE → yfinance
# ─────────────────────────────────────────────────────────
def fetch_via_vci(ticker, start_date, end_date):
    stock = Vnstock().stock(symbol=ticker, source='VCI')
    df = stock.quote.history(start=start_date, end=end_date, interval='1D')
    if df is None or df.empty:
        return None
    df.columns = [str(c).lower().strip() for c in df.columns]
    df['_source'] = 'VCI'
    return df

def fetch_via_kb(ticker, start_date, end_date):
    """KB Securities REST API"""
    url = "https://api.kbsec.com/stock/historyprice"
    params = {
        "symbol"   : ticker,
        "fromDate" : start_date.replace("-", ""),   # YYYYMMDD
        "toDate"   : end_date.replace("-", ""),
        "period"   : "D"
    }
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, params=params, headers=headers, timeout=10)
    r.raise_for_status()
    data = r.json()

    # KB trả về list dict hoặc {"data": [...]}
    rows = data if isinstance(data, list) else data.get("data", [])
    if not rows:
        return None

    df = pd.DataFrame(rows)
    df.columns = [c.lower().strip() for c in df.columns]

    # Chuẩn hóa tên cột KB → chuẩn
    rename = {
        "tradingdate": "time", "date": "time",
        "openprice": "open",   "highprice": "high",
        "lowprice": "low",     "closeprice": "close",
        "totalvolume": "volume"
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    df['_source'] = 'KB'
    return df

def fetch_via_dnse(ticker, start_date, end_date):
    """DNSE Entrade API"""
    url = f"https://services.entrade.com.vn/chart-api/v2/ohlcs/stock"
    params = {
        "symbol"  : ticker,
        "from"    : start_date,
        "to"      : end_date,
        "resolution": "D"
    }
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, params=params, headers=headers, timeout=10)
    r.raise_for_status()
    data = r.json()

    # DNSE trả về {"t":[], "o":[], "h":[], "l":[], "c":[], "v":[]}
    if not data.get("t"):
        return None

    df = pd.DataFrame({
        "time"  : pd.to_datetime(data["t"], unit="s").strftime("%Y-%m-%d"),
        "open"  : data["o"],
        "high"  : data["h"],
        "low"   : data["l"],
        "close" : data["c"],
        "volume": data["v"],
    })
    df['_source'] = 'DNSE'
    return df

def fetch_via_yfinance(ticker, exchange, start_date, end_date):
    suffix = YF_SUFFIX.get(exchange, ".VN")
    df = yf.download(f"{ticker}{suffix}", start=start_date, end=end_date,
                     progress=False, auto_adjust=True)
    if df is None or df.empty:
        return None
    df = df.reset_index()
    df.columns = [str(c).lower().strip() for c in df.columns]
    df = df.rename(columns={"date": "time", "adj close": "close"})
    df['_source'] = 'yfinance'
    return df

def parse_row(df, ticker):
    row = df.iloc[-1]
    close = float(row['close'])
    open_ = float(row.get('open', close))
    high  = float(row.get('high', close))
    low   = float(row.get('low',  close))

    if close < 1000:
        close *= 1000; open_ *= 1000; high *= 1000; low *= 1000

    pct = round((close - open_) / open_ * 100, 2) if open_ else 0

    return {
        "Mã CK"      : ticker,
        "Ngày"       : str(row.get('time', '')),
        "Mở Cửa"     : int(open_),
        "Cao Nhất"   : int(high),
        "Thấp Nhất"  : int(low),
        "Đóng Cửa"   : int(close),
        "Khối Lượng" : int(float(row.get('volume', 0))),
        "% Thay Đổi" : pct,
        "Nguồn"      : str(row.get('_source', '')),
    }

def fetch_latest_price(ticker, exchange, start_date, end_date):
    # 1. VCI
    try:
        df = fetch_via_vci(ticker, start_date, end_date)
        if df is not None:
            return parse_row(df, ticker)
    except Exception:
        pass

    # 2. KB
    try:
        df = fetch_via_kb(ticker, start_date, end_date)
        if df is not None:
            return parse_row(df, ticker)
    except Exception:
        pass

    # 3. DNSE
    try:
        df = fetch_via_dnse(ticker, start_date, end_date)
        if df is not None:
            return parse_row(df, ticker)
    except Exception:
        pass

    # 4. yfinance
    try:
        df = fetch_via_yfinance(ticker, exchange, start_date, end_date)
        if df is not None:
            return parse_row(df, ticker)
    except Exception:
        pass

    return None

# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────
def main():
    today = datetime.now().strftime('%Y-%m-%d')
    start = (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d')
    print(f"📅 Export ngày: {today}\n")

    print("📋 Lấy danh sách mã theo sàn...")
    exchange_map = get_tickers_by_exchange()

    all_data = {}
    for exchange, tickers in exchange_map.items():
        print(f"\n🔄 {exchange} ({len(tickers)} mã)...")
        rows = []
        for ticker in tickers:
            result = fetch_latest_price(ticker, exchange, start, today)
            if result:
                rows.append(result)
            time.sleep(0.3)
        all_data[exchange] = pd.DataFrame(rows) if rows else pd.DataFrame()

        # Thống kê nguồn dữ liệu
        if rows:
            df_src = all_data[exchange]['Nguồn'].value_counts()
            print(f"  ✅ {len(rows)} mã — {df_src.to_dict()}")

    print("\n📁 Xuất file Excel...")
    export_xlsx(all_data)

    print("\n📤 Đẩy lên Google Sheet...")
    push_to_gsheet(all_data)

    print("\n🎉 Hoàn thành!")

if __name__ == "__main__":
    main()
